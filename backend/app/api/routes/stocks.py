"""
Stock API Routes
Endpoints untuk:
  - Upload file MB52 & ZSD_SODO dari SAP
  - Query hasil kalkulasi stok
  - Preview perhitungan stok
  - Recalculate stok
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date
from typing import List, Optional
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.core.database import get_db
from app.core.time import to_utc_iso
from app.models.models import StockCalculation, SAPUpload, Warehouse, WarehousePlant
from app.schemas.schemas import StockCalculationOut, StockPreview, SAPUploadOut, UploadResponse
from app.services.sap_parser import parse_mb52, parse_zsd_sodo, preview_mb52, preview_zsd_sodo
from app.services.stock_calculator import (
    calculate_daily_stock,
    update_stock_calculation,
    recalculate_all_for_date,
    save_mb52_data,
    save_zsd_sodo_data,
)

router = APIRouter()


# ==================== Upload Endpoints ====================

@router.post("/upload/mb52", response_model=UploadResponse)
async def upload_mb52(
    file: UploadFile = File(...),
    tanggal: date = Form(...),
    db: Session = Depends(get_db),
):
    """
    Upload file MB52 dari SAP.
    File akan di-parse, data stok mentah disimpan ke tabel sap_stock,
    lalu kalkulasi stok otomatis dijalankan untuk semua plant pada tanggal tersebut.
    """
    fname = (file.filename or "").lower()
    if not fname.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail=f"File harus berformat .xlsx atau .xls (diterima: {file.filename})")

    upload_record = SAPUpload(
        jenis_file="MB52",
        tanggal_data=tanggal,
        filename=file.filename,
        status="processing",
    )
    db.add(upload_record)
    db.commit()
    db.refresh(upload_record)

    try:
        file_bytes = await file.read()
        rows = parse_mb52(file_bytes, tanggal)
        
        if not rows:
            raise ValueError("Tidak ada data valid yang ditemukan dalam file MB52.")
        
        count = save_mb52_data(db, tanggal, rows)
        calc_results = recalculate_all_for_date(db, tanggal)
        
        upload_record.jumlah_baris = count
        upload_record.status = "success"
        db.commit()
        db.refresh(upload_record)
        
        return UploadResponse(
            message=f"File MB52 berhasil diproses. {count} baris data stok disimpan.",
            upload=SAPUploadOut.model_validate(upload_record),
            rows_processed=count,
            calculations_updated=len(calc_results),
        )

    except Exception as e:
        upload_record.status = "failed"
        upload_record.error_message = str(e)
        db.commit()
        raise HTTPException(status_code=400, detail=f"Gagal memproses file MB52: {str(e)}")


@router.post("/upload/zsd-sodo", response_model=UploadResponse)
async def upload_zsd_sodo(
    file: UploadFile = File(...),
    tanggal: date = Form(...),
    db: Session = Depends(get_db),
):
    """
    Upload file ZSD_SODO dari SAP.
    File akan di-parse, data outstanding DO disimpan ke tabel sap_outstanding_do,
    lalu kalkulasi stok otomatis dijalankan untuk semua plant pada tanggal tersebut.
    """
    fname = (file.filename or "").lower()
    if not fname.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail=f"File harus berformat .xlsx atau .xls (diterima: {file.filename})")

    upload_record = SAPUpload(
        jenis_file="ZSD_SODO",
        tanggal_data=tanggal,
        filename=file.filename,
        status="processing",
    )
    db.add(upload_record)
    db.commit()
    db.refresh(upload_record)

    try:
        file_bytes = await file.read()
        rows = parse_zsd_sodo(file_bytes, tanggal)
        
        if not rows:
            raise ValueError("Tidak ada data valid yang ditemukan dalam file ZSD_SODO.")
        
        count = save_zsd_sodo_data(db, tanggal, rows)
        calc_results = recalculate_all_for_date(db, tanggal)
        
        upload_record.jumlah_baris = count
        upload_record.status = "success"
        db.commit()
        db.refresh(upload_record)
        
        return UploadResponse(
            message=f"File ZSD_SODO berhasil diproses. {count} baris outstanding DO disimpan.",
            upload=SAPUploadOut.model_validate(upload_record),
            rows_processed=count,
            calculations_updated=len(calc_results),
        )

    except Exception as e:
        upload_record.status = "failed"
        upload_record.error_message = str(e)
        db.commit()
        raise HTTPException(status_code=400, detail=f"Gagal memproses file ZSD_SODO: {str(e)}")


@router.post("/upload/mb52/dry-run")
async def dry_run_mb52(
    file: UploadFile = File(...),
    tanggal: date = Form(...),
):
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail=f"File harus berformat .xlsx atau .xls (diterima: {file.filename})")
    file_bytes = await file.read()
    try:
        return preview_mb52(file_bytes, tanggal)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Dry-run MB52 gagal: {str(e)}")


@router.post("/upload/zsd-sodo/dry-run")
async def dry_run_zsd_sodo(
    file: UploadFile = File(...),
    tanggal: date = Form(...),
):
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail=f"File harus berformat .xlsx atau .xls (diterima: {file.filename})")
    file_bytes = await file.read()
    try:
        return preview_zsd_sodo(file_bytes, tanggal)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Dry-run ZSD_SODO gagal: {str(e)}")


# ==================== Query Endpoints ====================

@router.get("/", response_model=List[StockCalculationOut])
def get_stocks(
    db: Session = Depends(get_db),
    tanggal: Optional[date] = None,
    gudang_id: Optional[int] = None,
    tipe_pupuk: Optional[str] = None,
):
    """Ambil daftar hasil kalkulasi stok dengan filter opsional."""
    query = db.query(StockCalculation)
    
    if not tanggal:
        # Get the latest date available in the database
        latest_date = db.query(func.max(StockCalculation.tanggal)).scalar()
        tanggal = latest_date

    if tanggal:
        query = query.filter(StockCalculation.tanggal == tanggal)
    if gudang_id:
        query = query.filter(StockCalculation.gudang_id == gudang_id)
    if tipe_pupuk:
        query = query.filter(StockCalculation.tipe_pupuk == tipe_pupuk)
    
    query = query.order_by(StockCalculation.gudang_id, StockCalculation.tipe_pupuk)
    
    return query.all()


@router.get("/intransit-range")
def get_intransit_range(
    gudang_id: int,
    tanggal_awal: date,
    tanggal_akhir: date,
    db: Session = Depends(get_db),
):
    """Ambil data intransit per gudang dan rentang tanggal."""
    if tanggal_akhir < tanggal_awal:
        raise HTTPException(status_code=400, detail="Tanggal akhir tidak boleh lebih awal dari tanggal awal.")

    rows = (
        db.query(StockCalculation)
        .filter(
            StockCalculation.gudang_id == gudang_id,
            StockCalculation.tanggal.between(tanggal_awal, tanggal_akhir),
        )
        .order_by(StockCalculation.tipe_pupuk, StockCalculation.tanggal)
        .all()
    )

    return [
        {
            "tanggal": row.tanggal.isoformat(),
            "gudang_id": row.gudang_id,
            "nama_gudang": row.nama_gudang,
            "kode_plants": row.kode_plants,
            "kota": row.kota,
            "provinsi": row.provinsi,
            "tipe_pupuk": row.tipe_pupuk,
            "intransit": row.intransit,
        }
        for row in rows
    ]


@router.get("/preview", response_model=StockPreview)
def preview_calculation(
    tanggal: date,
    gudang_id: int,
    tipe_pupuk: str,
    db: Session = Depends(get_db),
):
    """Preview hasil hitungan tanpa menyimpan ke database."""
    try:
        calc_data = calculate_daily_stock(db, tanggal, gudang_id, tipe_pupuk)
        return calc_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/recalculate")
def recalculate_stocks(
    tanggal: date = Form(...),
    db: Session = Depends(get_db),
):
    """Hitung ulang semua stok untuk tanggal tertentu."""
    try:
        results = recalculate_all_for_date(db, tanggal)
        return {
            "message": f"Berhasil menghitung ulang stok untuk {len(results)} kombinasi gudang/pupuk.",
            "tanggal": tanggal.isoformat(),
            "calculations": results,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== Upload History ====================

def _upload_status_payload(upload: Optional[SAPUpload]):
    return {
        "uploaded": upload is not None,
        "latest_upload_at": to_utc_iso(upload.created_at) if upload and upload.created_at else None,
        "filename": upload.filename if upload else None,
        "rows": upload.jumlah_baris if upload else 0,
    }


@router.get("/upload-status")
def get_upload_status(
    tanggal: date,
    db: Session = Depends(get_db),
):
    """Cek kelengkapan upload MB52 dan ZSD_SODO untuk satu tanggal data."""
    uploads = (
        db.query(SAPUpload)
        .filter(
            SAPUpload.tanggal_data == tanggal,
            SAPUpload.status == "success",
            SAPUpload.jenis_file.in_(["MB52", "ZSD_SODO"]),
        )
        .order_by(SAPUpload.created_at.desc())
        .all()
    )

    latest_by_type: dict[str, SAPUpload] = {}
    for upload in uploads:
        if upload.jenis_file not in latest_by_type:
            latest_by_type[upload.jenis_file] = upload

    mb52 = latest_by_type.get("MB52")
    zsd_sodo = latest_by_type.get("ZSD_SODO")
    missing = []
    if not mb52:
        missing.append("MB52")
    if not zsd_sodo:
        missing.append("ZSD_SODO")

    if not missing:
        message = "Data lengkap: MB52 dan ZSD_SODO sudah diupload."
    elif len(missing) == 2:
        message = "Belum ada upload MB52 dan ZSD_SODO untuk tanggal ini."
    else:
        message = f"Data belum lengkap: {missing[0]} belum diupload."

    return {
        "tanggal": tanggal.isoformat(),
        "complete": len(missing) == 0,
        "mb52": _upload_status_payload(mb52),
        "zsd_sodo": _upload_status_payload(zsd_sodo),
        "missing": missing,
        "message": message,
    }


@router.get("/uploads", response_model=List[SAPUploadOut])
def get_upload_history(
    db: Session = Depends(get_db),
    jenis_file: Optional[str] = None,
    limit: int = 20,
):
    """Ambil riwayat upload file SAP."""
    query = db.query(SAPUpload).order_by(SAPUpload.created_at.desc()).limit(limit)
    
    if jenis_file:
        query = query.filter(SAPUpload.jenis_file == jenis_file)
    
    return query.all()


PROVINCE_MAPPING = {
    # Sumatera (termasuk Lampung)
    "aceh": "Sumatera",
    "sumatera utara": "Sumatera",
    "sumatera barat": "Sumatera",
    "riau": "Sumatera",
    "kepulauan riau": "Sumatera",
    "jambi": "Sumatera",
    "bengkulu": "Sumatera",
    "sumatera selatan": "Sumatera",
    "kepulauan bangka belitung": "Sumatera",
    "bangka belitung": "Sumatera",
    "lampung": "Sumatera",
    
    # Jawa
    "banten": "Jawa",
    "dki jakarta": "Jawa",
    "dki": "Jawa",
    "jakarta": "Jawa",
    "jawa barat": "Jawa",
    "jawa tengah": "Jawa",
    "di yogyakarta": "Jawa",
    "yogyakarta": "Jawa",
    "jawa timur": "Jawa",
    
    # Bali & Nusa Tenggara
    "bali": "Bali & NT",
    "nusa tenggara barat": "Bali & NT",
    "ntb": "Bali & NT",
    "nusa tenggara timur": "Bali & NT",
    "ntt": "Bali & NT",
    
    # Kalimantan
    "kalimantan barat": "Kalimantan",
    "kalimantan tengah": "Kalimantan",
    "kalimantan selatan": "Kalimantan",
    "kalimantan timur": "Kalimantan",
    "kalimantan utara": "Kalimantan",
}

def get_area_for_province(province: str) -> str:
    if not province:
        return "Sulamapa"
    prov_lower = province.strip().lower()
    for key, area in PROVINCE_MAPPING.items():
        if key in prov_lower:
            return area
    return "Sulamapa"

@router.get("/export-excel")
def export_excel(
    tanggal: date,
    db: Session = Depends(get_db),
):
    # Fetch all active warehouses
    warehouses = db.query(Warehouse).filter(Warehouse.is_active == True).all()
    
    # Fetch all calculations for the target date
    calcs = db.query(StockCalculation).filter(StockCalculation.tanggal == tanggal).all()
    
    # Map calculations for fast lookup: (gudang_id, tipe_pupuk) -> calculation
    calc_map = {}
    for c in calcs:
        calc_map[(c.gudang_id, c.tipe_pupuk)] = c

    # Areas defined by user
    areas = ["Sumatera", "Jawa", "Bali & NT", "Kalimantan", "Sulamapa"]
    fertilizers = ["Urea", "NPK"]

    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet:
        wb.remove(default_sheet)

    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "Kabupaten / Kota",
        "Kode Plant",
        "Nama Gudang",
        "Stok Fisik",
        "Outstanding SO",
        "Stok Admin (Tanpa Intransit)",
        "Intransit",
        "Stok Admin"
    ]

    for area in areas:
        for fert in fertilizers:
            # Filter warehouses in this area
            area_warehouses = []
            for w in warehouses:
                w_area = get_area_for_province(w.provinsi)
                if w_area == area:
                    area_warehouses.append(w)
            
            # Group by province -> city/kabupaten
            prov_groups = {}
            for w in area_warehouses:
                prov = (w.provinsi or "TIDAK DIKETAHUI").strip().upper()
                city = (w.kota or "TIDAK DIKETAHUI").strip().upper()
                if prov not in prov_groups:
                    prov_groups[prov] = {}
                if city not in prov_groups[prov]:
                    prov_groups[prov][city] = []
                prov_groups[prov][city].append(w)

            # Sort provinces by minimum province code (first 2 digits of kode_kab)
            def get_province_sort_key(prov_name):
                all_w = []
                for city in prov_groups[prov_name].values():
                    all_w.extend(city)
                codes = []
                for w in all_w:
                    if w.kode_kab:
                        try:
                            codes.append(int(w.kode_kab) // 100)
                        except (ValueError, TypeError):
                            pass
                return min(codes) if codes else 99

            sorted_provinces = sorted(prov_groups.keys(), key=lambda p: (get_province_sort_key(p), p))

            # Sheet name
            sheet_title = f"{area} - {fert}"
            ws = wb.create_sheet(title=sheet_title)
            
            # Show grid lines
            ws.views.sheetView[0].showGridLines = True

            # Write Title
            ws.cell(row=1, column=1, value=f"LAPORAN HARIAN STOK {fert.upper()} - WILAYAH {area.upper()}").font = Font(name="Calibri", size=14, bold=True)
            ws.cell(row=2, column=1, value=f"Tanggal: {tanggal.strftime('%d %B %Y')}").font = Font(name="Calibri", size=11, italic=True)
            
            # Write Headers
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            ws.row_dimensions[4].height = 28
            
            current_row = 5
            total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            total_font = Font(name="Calibri", size=11, bold=True)

            for prov in sorted_provinces:
                # Accumulators for province totals
                prov_stok_fisik = 0.0
                prov_outstanding_so = 0.0
                prov_stok_admin_tanpa_intransit = 0.0
                prov_intransit = 0.0
                prov_stok_admin = 0.0

                sorted_cities = sorted(prov_groups[prov].keys())
                for city in sorted_cities:
                    gudangs = sorted(prov_groups[prov][city], key=lambda x: (x.nama_gudang or "").lower())
                    start_row = current_row
                    
                    for w in gudangs:
                        # Fetch calculation
                        calc = calc_map.get((w.id, fert))
                        stok_fisik = calc.stok_fisik if calc else 0.0
                        outstanding_so = calc.outstanding_so if calc else 0.0
                        stok_admin_tanpa_intransit = calc.stok_admin_tanpa_intransit if calc else 0.0
                        intransit = calc.intransit if calc else 0.0
                        stok_admin = calc.stok_admin if calc else 0.0
                        
                        # Add to province totals
                        prov_stok_fisik += stok_fisik
                        prov_outstanding_so += outstanding_so
                        prov_stok_admin_tanpa_intransit += stok_admin_tanpa_intransit
                        prov_intransit += intransit
                        prov_stok_admin += stok_admin

                        kode_plants = "/".join(sorted([p.kode_plant for p in w.plants]))
                        
                        ws.cell(row=current_row, column=1, value=city).alignment = center_align
                        ws.cell(row=current_row, column=2, value=kode_plants).alignment = center_align
                        ws.cell(row=current_row, column=3, value=w.nama_gudang).alignment = left_align
                        
                        # Numeric values
                        for col_idx, val in enumerate([stok_fisik, outstanding_so, stok_admin_tanpa_intransit, intransit, stok_admin], 4):
                            cell = ws.cell(row=current_row, column=col_idx, value=val)
                            cell.alignment = right_align
                            cell.number_format = '#,##0.00'
                        
                        # Apply font & border to row
                        for col_idx in range(1, 9):
                            cell = ws.cell(row=current_row, column=col_idx)
                            cell.font = regular_font
                            cell.border = thin_border
                        
                        current_row += 1
                    
                    end_row = current_row - 1
                    if end_row >= start_row:
                        if end_row > start_row:
                            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                        # Force center alignment on merged area
                        ws.cell(row=start_row, column=1).alignment = center_align

                # Write Province Total Row
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
                total_label_cell = ws.cell(row=current_row, column=1, value=f"Total Stok Provinsi {prov.title()}")
                total_label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                
                # Write sums
                ws.cell(row=current_row, column=4, value=prov_stok_fisik)
                ws.cell(row=current_row, column=5, value=prov_outstanding_so)
                ws.cell(row=current_row, column=6, value=prov_stok_admin_tanpa_intransit)
                ws.cell(row=current_row, column=7, value=prov_intransit)
                ws.cell(row=current_row, column=8, value=prov_stok_admin)

                # Format total row
                ws.row_dimensions[current_row].height = 22
                for col_idx in range(1, 9):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font = total_font
                    cell.fill = total_fill
                    cell.border = thin_border
                    if col_idx >= 4:
                        cell.alignment = right_align
                        cell.number_format = '#,##0.00'
                
                current_row += 1

            # Auto-fit columns
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.row < 4:
                        continue  # Skip title row lengths
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to IO stream
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"Laporan_Stok_{tanggal.isoformat()}.xlsx"
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

